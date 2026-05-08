# ===== GPC Orders (cloud, GitHub storage, st.secrets/SEC) =====
import os
from io import BytesIO, StringIO
from datetime import datetime
import base64
import json
import requests
from typing import Optional

import pandas as pd
import streamlit as st
import yaml

import hashlib
import hmac
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import streamlit.components.v1 as components

SEC = dict(st.secrets)

st.set_page_config(page_title="GPC Orders Systeem", layout="wide")

HERE = os.path.dirname(__file__)
AUTH_YAML = os.path.join(HERE, "auth.yaml")


def _gh_headers():
    return {
        "Authorization": f"Bearer {SEC['GITHUB_TOKEN']}",
        "Accept": "application/vnd.github+json",
    }

def _gh_api(path: str) -> str:
    owner = SEC["GITHUB_OWNER"]
    repo = SEC["GITHUB_REPO"]
    return f"https://api.github.com/repos/{owner}/{repo}{path}"

def _gh_get_text(path_in_repo: str) -> Optional[str]:
    url = _gh_api(f"/contents/{path_in_repo}")
    try:
        r = requests.get(url, headers=_gh_headers(), timeout=10)
    except Exception as e:
        st.error(f"GitHub verbinding mislukt: {e}")
        return ""

    if r.status_code == 200:
        data = r.json()
        return base64.b64decode(data["content"]).decode("utf-8", errors="ignore")

    if r.status_code == 404:
        return None

    st.error(f"GitHub leesfout {r.status_code}: {r.text[:200]}")
    return ""

def _gh_put_text(path_in_repo: str, content_text: str, msg: str):
    url = _gh_api(f"/contents/{path_in_repo}")

    try:
        r = requests.get(url, headers=_gh_headers(), timeout=10)
    except Exception as e:
        st.error(f"GitHub schrijf-verbinding mislukt (voor lezen sha): {e}")
        return

    sha = r.json().get("sha") if r.status_code == 200 else None

    payload = {
        "message": msg,
        "content": base64.b64encode(content_text.encode("utf-8")).decode("ascii"),
        "branch": "main",
    }

    if sha:
        payload["sha"] = sha

    try:
        r2 = requests.put(url, headers=_gh_headers(), data=json.dumps(payload), timeout=10)
    except Exception as e:
        st.error(f"GitHub schrijf-verbinding mislukt: {e}")
        return

    if r2.status_code not in (200, 201):
        st.error(f"GitHub schrijffout {r2.status_code}: {r2.text[:200]}")

def _gh_get_csv(path_in_repo: str) -> Optional[pd.DataFrame]:
    txt = _gh_get_text(path_in_repo)

    if txt is None:
        return None

    if not txt.strip():
        return pd.DataFrame()

    try:
        return pd.read_csv(StringIO(txt))
    except Exception:
        return pd.DataFrame()

def _gh_put_csv(path_in_repo: str, df: pd.DataFrame, msg: str):
    csv_txt = df.to_csv(index=False)
    _gh_put_text(path_in_repo, csv_txt, msg)


def load_auth() -> dict:
    try:
        with open(AUTH_YAML, "r", encoding="utf-8") as f:
            return yaml.safe_load(f) or {}
    except Exception:
        st.error("auth.yaml niet gevonden of ongeldig. Voeg auth.yaml toe aan de repo.")
        return {}

def login_panel():
    cfg = load_auth()
    users = cfg.get("credentials", {}).get("usernames", {})
    st.session_state.setdefault("auth_user", None)

    if st.session_state["auth_user"]:
        return st.session_state["auth_user"]

    st.markdown("### 🔐 Inloggen")
    u = st.text_input("Gebruikersnaam")
    p = st.text_input("Wachtwoord", type="password")

    if st.button("Inloggen", type="primary"):
        rec = users.get(u)

        if rec:
            ok = False

            if isinstance(rec.get("password_sha256"), str) and rec["password_sha256"]:
                try:
                    entered = hashlib.sha256(str(p).encode("utf-8")).hexdigest()
                    ok = hmac.compare_digest(entered, rec["password_sha256"])
                except Exception:
                    ok = False

            elif "password_plain" in rec:
                ok = str(p) == str(rec["password_plain"])

            if ok:
                st.session_state["auth_user"] = {
                    "username": u,
                    "name": rec.get("name", u),
                    "email": rec.get("email", "")
                }
                st.success(f"Ingelogd als {st.session_state['auth_user']['name']}")
                st.rerun()
            else:
                st.error("Ongeldige gebruikersnaam/wachtwoord")
        else:
            st.error("Ongeldige gebruikersnaam/wachtwoord")

    st.stop()


def week_start_date(year: int, week: int):
    try:
        return datetime.fromisocalendar(int(year), int(week), 1).date()
    except Exception:
        return None

def coerce_columns(df: pd.DataFrame, types: dict) -> pd.DataFrame:
    df = df.copy()

    for col, dtype in types.items():
        if col not in df.columns:
            if dtype in ("Int64", "Int32", "int"):
                df[col] = pd.Series([], dtype="Int64")
            elif dtype in ("float", "Float64"):
                df[col] = pd.Series([], dtype="float")
            else:
                df[col] = pd.Series([], dtype="string")
        else:
            try:
                if dtype in ("Int64", "Int32", "int"):
                    df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64")
                elif dtype in ("float", "Float64"):
                    df[col] = pd.to_numeric(df[col], errors="coerce").astype(float)
                else:
                    df[col] = df[col].astype("string")
            except Exception:
                pass

    return df

def load_data():
    repo_dir = SEC.get("DATA_DIR", "data")

    g = _gh_get_csv(f"{repo_dir}/products.csv")
    prod = pd.DataFrame(columns=["id", "name", "description", "price", "four_week_availability", "supplier"]) if g is None else g
    prod = coerce_columns(prod, {
        "id": "Int64",
        "name": "string",
        "description": "string",
        "price": "float",
        "four_week_availability": "Int64",
        "supplier": "string",
    })
    st.session_state.products = prod

    g = _gh_get_csv(f"{repo_dir}/customers.csv")
    cust = pd.DataFrame(columns=["id", "name", "email"]) if g is None else g
    cust = coerce_columns(cust, {
        "id": "Int64",
        "name": "string",
        "email": "string",
    })
    st.session_state.customers = cust

    g = _gh_get_csv(f"{repo_dir}/orders.csv")
    ords = pd.DataFrame(columns=[
        "id", "customer_id", "product_id", "quantity", "week_number", "year", "notes", "sales_price"
    ]) if g is None else g
    ords = coerce_columns(ords, {
        "id": "Int64",
        "customer_id": "Int64",
        "product_id": "Int64",
        "quantity": "Int64",
        "week_number": "Int64",
        "year": "Int64",
        "notes": "string",
        "sales_price": "float",
    })
    st.session_state.orders = ords

def save_data():
    repo_dir = SEC.get("DATA_DIR", "data")

    if "products" in st.session_state:
        _gh_put_csv(f"{repo_dir}/products.csv", st.session_state.products, "update products.csv")

    if "customers" in st.session_state:
        _gh_put_csv(f"{repo_dir}/customers.csv", st.session_state.customers, "update customers.csv")

    if "orders" in st.session_state:
        _gh_put_csv(f"{repo_dir}/orders.csv", st.session_state.orders, "update orders.csv")

def ensure_state():
    if "products" not in st.session_state or "customers" not in st.session_state or "orders" not in st.session_state:
        load_data()

    st.session_state.products = coerce_columns(st.session_state.products, {
        "id": "Int64",
        "name": "string",
        "description": "string",
        "price": "float",
        "four_week_availability": "Int64",
        "supplier": "string",
    })

    st.session_state.customers = coerce_columns(st.session_state.customers, {
        "id": "Int64",
        "name": "string",
        "email": "string",
    })

    st.session_state.orders = coerce_columns(st.session_state.orders, {
        "id": "Int64",
        "customer_id": "Int64",
        "product_id": "Int64",
        "quantity": "Int64",
        "week_number": "Int64",
        "year": "Int64",
        "notes": "string",
        "sales_price": "float",
    })

def next_id(df: pd.DataFrame) -> int:
    if df.empty or "id" not in df.columns:
        return 1

    try:
        return int(pd.to_numeric(df["id"], errors="coerce").fillna(0).max()) + 1
    except Exception:
        return 1

def label_product_with_supplier(prod_id: Optional[int]) -> str:
    try:
        if prod_id is None:
            return ""

        pid = int(prod_id)
        df = st.session_state.products
        row = df.loc[pd.to_numeric(df["id"], errors="coerce") == pid]

        if row.empty:
            return ""

        name = str(row.iloc[0]["name"] or "")
        supplier = str(row.iloc[0]["supplier"] or "")
        return f"{name} — {supplier}" if supplier else name
    except Exception:
        return ""

def fmt_select_from_df(id_value, df_id_name: pd.DataFrame) -> str:
    try:
        if id_value is None:
            return ""

        iid = int(id_value)
        m = df_id_name.loc[pd.to_numeric(df_id_name["id"], errors="coerce") == iid, "name"]
        return "" if m.empty else str(m.iloc[0])
    except Exception:
        return ""

def enable_enter_navigation(submit_button_label: str):
    components.html(f"""
    <script>
    (function() {{
      const root = window.parent.document;
      function isEditable(el) {{
        if (!el) return false;
        const tag = el.tagName;
        if (tag === 'INPUT' || tag === 'TEXTAREA') return true;
        if (el.getAttribute && el.getAttribute('contenteditable') === 'true') return true;
        return false;
      }}
      function getFocusableInputs() {{
        const all = Array.from(root.querySelectorAll('input, textarea'));
        return all.filter(el => !el.disabled && el.offsetParent !== null);
      }}
      function findSubmitButton(label) {{
        const btns = Array.from(root.querySelectorAll('button'));
        return btns.find(b => (b.innerText || '').trim() === label.trim());
      }}
      function focusNext(current) {{
        const inputs = getFocusableInputs();
        const idx = inputs.indexOf(current);
        if (idx === -1) return false;
        const next = inputs[idx + 1];
        if (next) {{
          next.focus();
          if (next.setSelectionRange && next.value != null) {{
            const len = next.value.length;
            try {{ next.setSelectionRange(len, len); }} catch(e) {{}}
          }}
          return true;
        }}
        const btn = findSubmitButton("{submit_button_label}");
        if (btn) btn.click();
        return true;
      }}
      function isSelectOpenOrFocused(el) {{
        try {{
          const inCombobox = el && (el.closest('[data-baseweb="select"]') || el.closest('[role="combobox"]'));
          const popupOpen = root.querySelector('[role="listbox"]');
          return !!(inCombobox || popupOpen);
        }} catch(e) {{ return false; }}
      }}
      function handler(e) {{
        if (e.key !== 'Enter') return;
        const active = root.activeElement;
        if (isSelectOpenOrFocused(active)) return;
        if (isEditable(active)) {{
          e.preventDefault();
          focusNext(active);
        }}
      }}
      root.addEventListener('keydown', handler, true);
    }})();
    </script>
    """, height=0)

def money_input(label: str, value: float = 0.0, key: str = None, help: str = None):
    default_txt = f"{value:.2f}".replace(".", ",")
    raw = st.text_input(label, value=default_txt, key=key, help=help)
    txt = (raw or "").strip().replace("€", "").replace(" ", "").replace(",", ".")

    try:
        val = round(float(txt), 2)
        return val, True
    except Exception:
        return value, False


def build_orders_display_df() -> pd.DataFrame:
    orders = st.session_state.orders.copy()
    products = st.session_state.products.copy()
    customers = st.session_state.customers.copy()

    if orders.empty:
        return pd.DataFrame(columns=[
            "Customer", "Article", "Description", "Quantity", "Purchase Price", "Sales Price", "Supplier",
            "Week", "Week Start (Mon)", "Year", "_OID", "_CID", "_PID"
        ])

    if not products.empty:
        prod = products.rename(columns={"id": "_PID_join"})
        orders = orders.merge(
            prod[["_PID_join", "name", "description", "price", "supplier"]],
            left_on="product_id",
            right_on="_PID_join",
            how="left"
        )
    else:
        orders["_PID_join"] = None
        orders["name"] = ""
        orders["description"] = ""
        orders["price"] = None
        orders["supplier"] = ""

    if not customers.empty:
        cust = customers.rename(columns={"id": "_CID_join"})
        orders = orders.merge(
            cust[["_CID_join", "name"]],
            left_on="customer_id",
            right_on="_CID_join",
            how="left",
            suffixes=("", "_cust")
        )
        orders["Customer"] = orders["name_cust"].fillna("")
    else:
        orders["_CID_join"] = None
        orders["Customer"] = ""

    orders["Article"] = orders["name"].fillna("")
    orders["Description"] = orders["description"].fillna("")
    orders["Quantity"] = pd.to_numeric(orders["quantity"], errors="coerce").fillna(0).astype(int)
    orders["Purchase Price"] = pd.to_numeric(orders["price"], errors="coerce")
    orders["Sales Price"] = pd.to_numeric(orders["sales_price"], errors="coerce")
    orders["Supplier"] = orders["supplier"].astype("string").fillna("")
    orders["Week"] = pd.to_numeric(orders["week_number"], errors="coerce").fillna(0).astype(int)
    orders["Year"] = pd.to_numeric(orders["year"], errors="coerce").fillna(0).astype(int)
    orders["Week Start (Mon)"] = orders.apply(lambda r: week_start_date(r["Year"], r["Week"]), axis=1)

    orders["_OID"] = pd.to_numeric(orders["id"], errors="coerce").astype("Int64")
    orders["_CID"] = pd.to_numeric(orders["customer_id"], errors="coerce").astype("Int64")
    orders["_PID"] = pd.to_numeric(orders["product_id"], errors="coerce").astype("Int64")

    view_cols = [
        "Customer", "Article", "Description", "Quantity", "Purchase Price", "Sales Price", "Supplier",
        "Week", "Week Start (Mon)", "Year", "_OID", "_CID", "_PID"
    ]

    df = orders.reindex(columns=view_cols).copy()

    for c in ["Customer", "Article", "Description", "Supplier"]:
        df[c] = df[c].astype("string").fillna("")

    return df

def _excel_export_bytes(df: pd.DataFrame, title: str) -> BytesIO:
    df = df.copy().fillna("")
    wb = Workbook()
    ws = wb.active
    ws.title = "Export"

    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name="Aptos", bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    start_row = 3

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=str(col_name))
        cell.font = Font(name="Aptos", bold=True)
        cell.alignment = Alignment(vertical="center")

    for r_idx, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for c_idx, val in enumerate(row.tolist(), start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    data_rows = df.shape[0]
    first_row = start_row
    last_row = start_row + data_rows
    last_col = df.shape[1]

    if data_rows > 0 and last_col > 0:
        ref = f"A{first_row}:{get_column_letter(last_col)}{last_row}"
        tbl = Table(displayName="Table1", ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium11",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        ws.add_table(tbl)

    max_row = ws.max_row
    max_col = ws.max_column

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if cell.font and cell.font.bold:
                cell.font = Font(name="Aptos", bold=True, size=cell.font.sz or 11)
            else:
                cell.font = Font(name="Aptos", size=cell.font.sz or 11)

    for c in range(1, max_col + 1):
        col_letter = get_column_letter(c)
        values = [str(ws.cell(row=r, column=c).value or "") for r in range(1, max_row + 1)]
        max_len = max((len(v) for v in values), default=10)
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 35)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ------------------------------------------------------------
# Init state + Top menu
# ------------------------------------------------------------
user = login_panel()
ensure_state()

st.markdown("## 🌿 GPC Orders Systeem")

top_left, top_right = st.columns([5, 2])

with top_left:
    page = st.radio(
        "Navigatie",
        ["Dashboard", "Orders", "Klanten", "Producten"],
        horizontal=True,
        label_visibility="collapsed",
        key="top_navigation"
    )

with top_right:
    st.caption(f"👤 Ingelogd als **{user['name']}**")

    btn_save, btn_logout = st.columns(2)

    with btn_save:
        if st.button("💾 Opslaan", use_container_width=True):
            save_data()
            st.success("Opgeslagen.")

    with btn_logout:
        if st.button("Uitloggen", use_container_width=True):
            st.session_state["auth_user"] = None
            st.rerun()

st.markdown("---")


# ------------------------------------------------------------
# Dashboard
# ------------------------------------------------------------
if page == "Dashboard":
    st.title("📊 Dashboard")

    orders = st.session_state.orders
    products = st.session_state.products

    if orders.empty:
        st.info("Nog geen data. Voeg eerst producten/klanten/orders toe.")
    else:
        df = (
            orders.merge(products[["id", "name"]], left_on="product_id", right_on="id", how="left")
            .rename(columns={"name": "Product"})
            .drop(columns=["id_y"], errors="ignore")
        )

        jaren = sorted(df["year"].dropna().astype(int).unique().tolist())

        sel_jaar = st.selectbox(
            "Jaar",
            jaren,
            index=jaren.index(datetime.now().year) if datetime.now().year in jaren else 0
        )

        per_prod = (
            df[df["year"] == sel_jaar]
            .groupby("Product", dropna=False)["quantity"]
            .sum()
            .reset_index(name="Totaal verkocht")
            .sort_values("Totaal verkocht", ascending=False)
        )

        st.markdown(f"### Orders per product in {sel_jaar}")
        st.dataframe(per_prod, use_container_width=True)


# ------------------------------------------------------------
# Orders
# ------------------------------------------------------------
elif page == "Orders":
    st.title("📦 Orders")

    st.subheader("➕ Nieuwe order")

    if st.session_state.customers.empty or st.session_state.products.empty:
        st.warning("Je hebt klanten én producten nodig om een order toe te voegen.")
    else:
        with st.form("add_order_form", clear_on_submit=True):
            cA, cB = st.columns(2)

            with cA:
                cust_ids = st.session_state.customers["id"].dropna().astype(int).tolist()
                prod_ids = st.session_state.products["id"].dropna().astype(int).tolist()

                last_cust = st.session_state.get("last_customer_id")
                cust_options = [None] + cust_ids
                cust_index = cust_options.index(last_cust) if last_cust in cust_ids else 0

                sel_customer = st.selectbox(
                    "Klant *",
                    options=cust_options,
                    format_func=lambda i: "" if i is None else fmt_select_from_df(i, st.session_state.customers),
                    index=cust_index
                )

                sel_product = st.selectbox(
                    "Artikel (product) *",
                    options=[None] + prod_ids,
                    format_func=lambda i: "" if i is None else label_product_with_supplier(i),
                    index=0
                )

                amount = st.number_input("Aantal *", min_value=1, step=1, value=1)

            with cB:
                verkoop, sp_ok = money_input(
                    "Verkoopprijs (optioneel)",
                    value=0.00,
                    key="oi_sales_price",
                    help="Gebruik 12,34 of 12.34."
                )

                weeks_txt = st.text_input("Weeknummers * (komma gescheiden, bijv. 4,8,12)", value="")
                jaar = st.number_input("Jaar *", min_value=2020, max_value=2100, step=1, value=datetime.now().year)

            enable_enter_navigation("Order(s) toevoegen")
            submitted = st.form_submit_button("Order(s) toevoegen")

            if submitted:
                errors = []

                if sel_customer is None:
                    errors.append("Kies een klant.")

                if sel_product is None:
                    errors.append("Kies een product.")

                if not sp_ok:
                    errors.append("Verkoopprijs is ongeldig. Gebruik 12,34 of 12.34.")

                weken, bad = [], []

                if not weeks_txt.strip():
                    errors.append("Vul ten minste één weeknummer in.")
                else:
                    for p in [w.strip() for w in weeks_txt.split(",") if w.strip()]:
                        try:
                            w = int(p)
                            if 1 <= w <= 53:
                                weken.append(w)
                            else:
                                bad.append(p)
                        except Exception:
                            bad.append(p)

                weken = sorted(list(dict.fromkeys(weken)))

                if bad:
                    errors.append(f"Ongeldige weeknummers: {', '.join(bad)} (toegestaan: 1..53)")

                if errors:
                    for e in errors:
                        st.error(e)
                else:
                    base_id = next_id(st.session_state.orders)
                    rows = []

                    for idx, w in enumerate(weken):
                        rows.append({
                            "id": base_id + idx,
                            "customer_id": int(sel_customer),
                            "product_id": int(sel_product),
                            "quantity": int(amount),
                            "sales_price": float(verkoop) if verkoop is not None else None,
                            "week_number": int(w),
                            "year": int(jaar),
                        })

                    st.session_state.orders = pd.concat(
                        [st.session_state.orders, pd.DataFrame(rows)],
                        ignore_index=True
                    )

                    st.session_state["last_customer_id"] = int(sel_customer)
                    save_data()
                    st.success(f"Toegevoegd: {len(rows)} order(s) voor weken: {', '.join(map(str, weken))}")
                    st.rerun()

    st.markdown("---")

    base_df = build_orders_display_df()

    with st.expander("🔎 Filters (tabel & export)"):
        q = st.text_input(
            "Zoeken",
            value="",
            placeholder="Zoek in Customer, Supplier, Article of Description…"
        )

        f1, f2, f3, f4 = st.columns(4)

        with f1:
            flt_customer = st.multiselect(
                "Customer",
                options=sorted(base_df["Customer"].dropna().astype(str).unique().tolist())
            )

        with f2:
            flt_supplier = st.multiselect(
                "Supplier",
                options=sorted(base_df["Supplier"].dropna().astype(str).unique().tolist())
            )

        with f3:
            flt_article = st.multiselect(
                "Article",
                options=sorted(base_df["Article"].dropna().astype(str).unique().tolist())
            )

        with f4:
            unique_years = sorted(base_df["Year"].dropna().astype(int).unique().tolist())
            flt_years = st.multiselect("Year", options=unique_years)

            unique_weeks = sorted(base_df["Week"].dropna().astype(int).unique().tolist())
            flt_weeks = st.multiselect("Week", options=unique_weeks)

    filtered_df = base_df.copy()

    if q.strip():
        query = q.strip()
        mask = False
        for col in ["Customer", "Supplier", "Article", "Description"]:
            mask = mask | filtered_df[col].astype(str).str.contains(query, case=False, na=False)
        filtered_df = filtered_df[mask]

    if flt_customer:
        filtered_df = filtered_df[filtered_df["Customer"].isin(flt_customer)]

    if flt_supplier:
        filtered_df = filtered_df[filtered_df["Supplier"].isin(flt_supplier)]

    if flt_article:
        filtered_df = filtered_df[filtered_df["Article"].isin(flt_article)]

    if flt_years:
        filtered_df = filtered_df[filtered_df["Year"].isin(flt_years)]

    if flt_weeks:
        filtered_df = filtered_df[filtered_df["Week"].isin(flt_weeks)]

    if filtered_df.empty:
        st.info("Geen orders gevonden.")
    else:
        st.subheader("📋 Orders")

        editor_df = filtered_df.copy()
        editor_df.insert(0, "Selecteer", False)

        editor_df["Order ID"] = pd.to_numeric(editor_df["_OID"], errors="coerce").astype("Int64")
        editor_df["Customer ID"] = pd.to_numeric(editor_df["_CID"], errors="coerce").astype("Int64")
        editor_df["Product ID"] = pd.to_numeric(editor_df["_PID"], errors="coerce").astype("Int64")

        for c in ["Customer", "Article", "Description", "Supplier"]:
            editor_df[c] = editor_df[c].astype("string").fillna("")

        editor_df["Quantity"] = pd.to_numeric(editor_df["Quantity"], errors="coerce").fillna(0).astype(int)
        editor_df["Week"] = pd.to_numeric(editor_df["Week"], errors="coerce").fillna(0).astype(int)
        editor_df["Year"] = pd.to_numeric(editor_df["Year"], errors="coerce").fillna(0).astype(int)
        editor_df["Purchase Price"] = pd.to_numeric(editor_df["Purchase Price"], errors="coerce")
        editor_df["Sales Price"] = pd.to_numeric(editor_df["Sales Price"], errors="coerce")

        editor_df = editor_df[
            [
                "Selecteer",
                "Customer",
                "Article",
                "Description",
                "Quantity",
                "Purchase Price",
                "Sales Price",
                "Supplier",
                "Week",
                "Week Start (Mon)",
                "Year",
                "Order ID",
                "Customer ID",
                "Product ID",
            ]
        ].copy()

        edited_orders = st.data_editor(
            editor_df,
            use_container_width=True,
            hide_index=True,
            num_rows="fixed",
            column_config={
                "Selecteer": st.column_config.CheckboxColumn(),
                "Customer": st.column_config.TextColumn(disabled=True),
                "Article": st.column_config.TextColumn(disabled=True),
                "Description": st.column_config.TextColumn(disabled=True),
                "Quantity": st.column_config.NumberColumn(min_value=0, step=1),
                "Purchase Price": st.column_config.NumberColumn(format="%.4f"),
                "Sales Price": st.column_config.NumberColumn(format="%.2f"),
                "Supplier": st.column_config.TextColumn(disabled=True),
                "Week": st.column_config.NumberColumn(min_value=1, max_value=53, step=1),
                "Week Start (Mon)": st.column_config.DateColumn(disabled=True),
                "Year": st.column_config.NumberColumn(min_value=2020, max_value=2100, step=1),
                "Order ID": None,
                "Customer ID": None,
                "Product ID": None,
            },
            key="orders_data_editor_v1",
        )

        c1, c2, _ = st.columns([1, 1, 6])

        with c1:
            if st.button("🗑️ Verwijder geselecteerde orders", use_container_width=True):
                del_ids = edited_orders.loc[
                    edited_orders["Selecteer"] == True,
                    "Order ID"
                ].dropna().astype(int).tolist()

                if not del_ids:
                    st.warning("Selecteer eerst één of meer orders.")
                else:
                    st.session_state.orders = st.session_state.orders[
                        ~st.session_state.orders["id"].isin(del_ids)
                    ]
                    save_data()
                    st.success(f"Verwijderd: {del_ids}")
                    st.rerun()

        with c2:
            if st.button("💾 Wijzigingen opslaan", use_container_width=True):
                try:
                    orders_base = st.session_state.orders.copy().set_index("id")
                    products_base = st.session_state.products.copy().set_index("id")

                    for _, row in edited_orders.iterrows():
                        oid = row.get("Order ID")
                        pid = row.get("Product ID")

                        if pd.notna(oid):
                            oid = int(oid)

                            if oid in orders_base.index:
                                qty = pd.to_numeric(row.get("Quantity"), errors="coerce")
                                week = pd.to_numeric(row.get("Week"), errors="coerce")
                                year = pd.to_numeric(row.get("Year"), errors="coerce")
                                sales_price = pd.to_numeric(row.get("Sales Price"), errors="coerce")

                                orders_base.at[oid, "quantity"] = int(qty) if pd.notna(qty) else 0
                                orders_base.at[oid, "week_number"] = int(week) if pd.notna(week) else 0
                                orders_base.at[oid, "year"] = int(year) if pd.notna(year) else datetime.now().year
                                orders_base.at[oid, "sales_price"] = None if pd.isna(sales_price) else round(float(sales_price), 2)

                        if pd.notna(pid):
                            pid = int(pid)

                            if pid in products_base.index:
                                purchase_price = pd.to_numeric(row.get("Purchase Price"), errors="coerce")

                                if not pd.isna(purchase_price):
                                    products_base.at[pid, "price"] = round(float(purchase_price), 4)

                    st.session_state.orders = orders_base.reset_index()
                    st.session_state.products = products_base.reset_index()

                    save_data()
                    st.success("Wijzigingen opgeslagen.")
                    st.rerun()

                except Exception as e:
                    st.error(f"Opslaan mislukt: {e}")

        st.markdown("### ⬇️ Export Excel")

        export_base = filtered_df.copy()

        export_base["Week"] = pd.to_numeric(export_base.get("Week", pd.Series(dtype="Int64")), errors="coerce").astype("Int64")
        export_base["Year"] = pd.to_numeric(export_base.get("Year", pd.Series(dtype="Int64")), errors="coerce").astype("Int64")
        export_base["Quantity"] = pd.to_numeric(export_base.get("Quantity", pd.Series(dtype="float")), errors="coerce").fillna(0).astype(int)

        def _mk_yw(row):
            y = row.get("Year")
            w = row.get("Week")

            if pd.isna(y) or pd.isna(w):
                return None

            try:
                return f"{int(y)}{int(w):02d}"
            except Exception:
                return None

        export_base["YearWeek"] = export_base.apply(_mk_yw, axis=1)

        def pivot_by_yearweek(df_src: pd.DataFrame, row_fields: list) -> pd.DataFrame:
            if df_src.empty:
                return pd.DataFrame(columns=row_fields)

            need = [c for c in row_fields + ["YearWeek", "Quantity"] if c in df_src.columns]
            df = df_src[need].copy()

            if df.empty:
                return pd.DataFrame(columns=row_fields)

            pvt = df.pivot_table(
                index=row_fields,
                columns="YearWeek",
                values="Quantity",
                aggfunc="sum",
                dropna=False
            )

            if isinstance(pvt.columns, pd.MultiIndex):
                pvt.columns = [c[-1] for c in pvt.columns]

            cols = [c for c in pvt.columns if c is not None]

            try:
                cols_sorted = sorted(cols, key=lambda x: int(x))
            except Exception:
                cols_sorted = sorted(cols)

            pvt = pvt.reindex(columns=cols_sorted)
            pvt = pvt.astype("float").where(pd.notna(pvt), None)
            pvt = pvt.reset_index()

            yw_cols = [c for c in pvt.columns if c not in row_fields]

            if yw_cols:
                tmp = pd.DataFrame(pvt[yw_cols]).fillna(0).sum(axis=1)
                pvt = pvt[tmp > 0]

            for c in row_fields:
                if c in pvt.columns:
                    pvt[c] = pvt[c].astype("string").fillna("")

            return pvt

        cust_rows = ["Customer", "Article", "Description", "Sales Price"]
        sup_rows = ["Supplier", "Article", "Description", "Customer", "Purchase Price"]

        cust_df = export_base[[c for c in cust_rows + ["YearWeek", "Quantity"] if c in export_base.columns]].copy()
        sup_df = export_base[[c for c in sup_rows + ["YearWeek", "Quantity"] if c in export_base.columns]].copy()

        cust_pivot = pivot_by_yearweek(cust_df, cust_rows)
        sup_pivot = pivot_by_yearweek(sup_df, sup_rows)

        cust_disabled = cust_pivot.empty
        sup_disabled = sup_pivot.empty

        cust_file = _excel_export_bytes(cust_pivot, f"GPC Orders {datetime.now().year}") if not cust_disabled else None
        sup_file = _excel_export_bytes(sup_pivot, f"GPC Orders {datetime.now().year}") if not sup_disabled else None

        e1, e2 = st.columns(2)

        with e1:
            st.download_button(
                "⬇️ Export Excel Customer",
                data=cust_file.getvalue() if cust_file else b"",
                file_name=f"GPC_Orders_Customer_{datetime.now().year}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                disabled=cust_disabled
            )

        with e2:
            st.download_button(
                "⬇️ Export Excel Supplier",
                data=sup_file.getvalue() if sup_file else b"",
                file_name=f"GPC_Orders_Supplier_{datetime.now().year}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                disabled=sup_disabled
            )


# ------------------------------------------------------------
# Klanten
# ------------------------------------------------------------
elif page == "Klanten":
    st.title("👥 Klanten")

    st.subheader("➕ Nieuwe klant")

    with st.form("add_customer_form", clear_on_submit=True):
        c1, c2 = st.columns(2)

        with c1:
            name = st.text_input("Naam *")

        with c2:
            email = st.text_input("E-mail")

        enable_enter_navigation("Klant toevoegen")
        ok = st.form_submit_button("Klant toevoegen")

    if ok and name.strip():
        new_row = {
            "id": next_id(st.session_state.customers),
            "name": name.strip(),
            "email": email.strip()
        }

        st.session_state.customers = pd.concat(
            [st.session_state.customers, pd.DataFrame([new_row])],
            ignore_index=True
        )

        save_data()
        st.success(f"Klant '{name}' toegevoegd.")
        st.rerun()

    st.markdown("---")

    if st.session_state.customers.empty:
        st.info("Nog geen klanten.")
    else:
        view = st.session_state.customers.copy().rename(columns={
            "id": "ID",
            "name": "Naam",
            "email": "E-mail"
        })

        view.insert(0, "Selecteer", False)

        st.subheader("✏️ Bewerken & Verwijderen")

        edited = st.data_editor(
            view,
            use_container_width=True,
            hide_index=True,
            num_rows="dynamic",
            column_config={
                "Selecteer": st.column_config.CheckboxColumn(),
                "ID": st.column_config.NumberColumn(disabled=True),
                "Naam": st.column_config.TextColumn(),
                "E-mail": st.column_config.TextColumn(),
            },
            key="customers_editor_v17"
        )

        if st.button("💾 Wijzigingen opslaan (Klanten)"):
            try:
                to_save = edited.drop(columns=["Selecteer"]).rename(columns={
                    "ID": "id",
                    "Naam": "name",
                    "E-mail": "email"
                })

                to_save = coerce_columns(to_save, {
                    "id": "int",
                    "name": "str",
                    "email": "str"
                })

                st.session_state.customers = to_save
                save_data()
                st.success("Klant-wijzigingen opgeslagen.")
                st.rerun()

            except Exception as e:
                st.error(f"Opslaan mislukt: {e}")

        sel_ids = edited.loc[edited["Selecteer"] == True, "ID"].tolist()

        if st.button("🗑️ Verwijder geselecteerde klanten"):
            if not sel_ids:
                st.warning("Selecteer eerst één of meer klanten.")
            else:
                st.session_state.customers = st.session_state.customers[
                    ~st.session_state.customers["id"].isin(sel_ids)
                ]

                save_data()
                st.success(f"Verwijderd: {sel_ids}")
                st.rerun()


# ------------------------------------------------------------
# Producten
# ------------------------------------------------------------
elif page == "Producten":
    st.title("🪴 Producten")

    st.subheader("➕ Nieuw product")

    with st.form("add_product_form", clear_on_submit=True):
        c1, c2 = st.columns(2)

        with c1:
            name = st.text_input("Productnaam *")

        with c2:
            price, price_ok = money_input(
                "Inkoopprijs (€)",
                value=0.00,
                key="pi_price",
                help="Gebruik 12,34 of 12.34."
            )
            fourw = st.number_input("Beschikbaarheid (4 weken)", min_value=0, value=0, step=1)
            supplier = st.text_input("Leverancier *")
            description = st.text_area("Omschrijving")

        enable_enter_navigation("Product toevoegen")
        ok = st.form_submit_button("Product toevoegen")

    if ok:
        errs = []

        if not name.strip():
            errs.append("Vul een productnaam in.")

        if not supplier.strip():
            errs.append("Vul een leverancier in.")

        if not price_ok:
            errs.append("Inkoopprijs is ongeldig. Gebruik 12,34 of 12.34.")

        if errs:
            for e in errs:
                st.error(e)
        else:
            new_row = {
                "id": next_id(st.session_state.products),
                "name": name.strip(),
                "description": description.strip(),
                "price": float(price),
                "four_week_availability": int(fourw),
                "supplier": supplier.strip(),
            }

            st.session_state.products = pd.concat(
                [st.session_state.products, pd.DataFrame([new_row])],
                ignore_index=True
            )

            save_data()
            st.success(f"Product '{name.strip()}' toegevoegd.")
            st.rerun()

    st.markdown("---")

    with st.expander("🛟 Veilige bewerkmodus"):
        if st.session_state.products.empty:
            st.info("Geen producten om te bewerken.")
        else:
            _pv = st.session_state.products.copy()
            _pv = coerce_columns(_pv, {
                "id": "int",
                "name": "str",
                "description": "str",
                "price": "float",
                "four_week_availability": "int",
                "supplier": "str"
            })

            _pv["label"] = _pv.apply(lambda r: f"{r['name']} ({r['supplier']}) [ID {r['id']}]", axis=1)
            _labels = _pv["label"].tolist()
            _id_by_label = dict(zip(_pv["label"], _pv["id"]))

            sel_label = st.selectbox("Kies product", options=_labels)
            sel_id = _id_by_label.get(sel_label)

            if sel_id is not None:
                row = _pv.loc[_pv["id"] == sel_id].iloc[0]

                with st.form(f"safe_edit_product_{sel_id}", clear_on_submit=False):
                    c1, c2 = st.columns(2)

                    with c1:
                        new_name = st.text_input("Naam", value=row["name"])
                        new_supplier = st.text_input("Leverancier", value=row["supplier"])
                        new_fourw = st.number_input(
                            "Beschikbaarheid (4 weken)",
                            min_value=0,
                            step=1,
                            value=int(row["four_week_availability"])
                        )

                    with c2:
                        new_price, ok_price = money_input(
                            "Inkoopprijs (€)",
                            value=float(row["price"] or 0.0),
                            key=f"safep_price_{sel_id}",
                            help="Gebruik 12,34 of 12.34"
                        )
                        new_desc = st.text_area("Omschrijving", value=row["description"] or "")

                    submit_safe = st.form_submit_button("💾 Opslaan (veilige modus)")

                if submit_safe:
                    errs = []

                    if not new_name.strip():
                        errs.append("Naam mag niet leeg zijn.")

                    if not new_supplier.strip():
                        errs.append("Leverancier mag niet leeg zijn.")

                    if not ok_price:
                        errs.append("Inkoopprijs is ongeldig.")

                    if errs:
                        for e in errs:
                            st.error(e)
                    else:
                        base = st.session_state.products.copy()
                        idx = base.index[base["id"] == sel_id]

                        if len(idx) == 1:
                            i = idx[0]
                            base.at[i, "name"] = new_name.strip()
                            base.at[i, "supplier"] = new_supplier.strip()
                            base.at[i, "four_week_availability"] = int(new_fourw)
                            base.at[i, "description"] = (new_desc or "").strip()
                            base.at[i, "price"] = float(new_price)

                            st.session_state.products = base
                            save_data()
                            st.success("Product bijgewerkt en opgeslagen ✅")
                            st.rerun()
                        else:
                            st.error("Kon de rij niet uniek vinden op ID.")

    if st.session_state.products.empty:
        st.info("Nog geen producten.")
    else:
        prod_view = st.session_state.products.copy()
        prod_view = coerce_columns(prod_view, {
            "id": "int",
            "name": "str",
            "description": "str",
            "price": "float",
            "four_week_availability": "int",
            "supplier": "str"
        })

        prod_view = prod_view.rename(columns={
            "id": "ID",
            "name": "Naam",
            "description": "Omschrijving",
            "price": "Inkoopprijs",
            "four_week_availability": "Beschikbaarheid (4w)",
            "supplier": "Leverancier"
        })

        prod_view.insert(0, "Selecteer", False)
        prod_view["ID"] = pd.to_numeric(prod_view["ID"], errors="coerce").fillna(0).astype(int)
        prod_view["Beschikbaarheid (4w)"] = pd.to_numeric(prod_view["Beschikbaarheid (4w)"], errors="coerce").fillna(0).astype(int)

        for _c in ["Naam", "Omschrijving", "Leverancier"]:
            prod_view[_c] = prod_view[_c].astype("string").fillna("")

        prod_view["Inkoopprijs"] = (
            pd.to_numeric(prod_view["Inkoopprijs"], errors="coerce")
            .apply(lambda v: "" if pd.isna(v) else f"{float(v):.2f}".replace(".", ","))
            .astype("string")
        )

        st.subheader("✏️ Bewerken & Verwijderen")

        edited = st.data_editor(
            prod_view,
            use_container_width=True,
            hide_index=True,
            num_rows="dynamic",
            column_config={
                "Selecteer": st.column_config.CheckboxColumn(),
                "ID": st.column_config.NumberColumn(disabled=True),
                "Naam": st.column_config.TextColumn(),
                "Omschrijving": st.column_config.TextColumn(),
                "Inkoopprijs": st.column_config.TextColumn(help="Gebruik 12,34 of 12.34"),
                "Beschikbaarheid (4w)": st.column_config.NumberColumn(format="%d", min_value=0, step=1),
                "Leverancier": st.column_config.TextColumn(),
            },
            key="product_editor_v17",
        )

        c1, c2 = st.columns(2)

        with c1:
            if st.button("💾 Wijzigingen opslaan (Producten)", use_container_width=True):
                try:
                    to_save = edited.drop(columns=["Selecteer"]).rename(columns={
                        "ID": "id",
                        "Naam": "name",
                        "Omschrijving": "description",
                        "Inkoopprijs": "price",
                        "Beschikbaarheid (4w)": "four_week_availability",
                        "Leverancier": "supplier"
                    })

                    if "price" in to_save.columns:
                        to_save["price"] = to_save["price"].astype(str).str.replace(",", ".", regex=False)
                        to_save["price"] = pd.to_numeric(to_save["price"], errors="coerce")

                    to_save = coerce_columns(to_save, {
                        "id": "int",
                        "name": "str",
                        "description": "str",
                        "price": "float",
                        "four_week_availability": "int",
                        "supplier": "str"
                    })

                    st.session_state.products = to_save
                    save_data()
                    st.success("Product-wijzigingen opgeslagen.")
                    st.rerun()

                except Exception as e:
                    st.error(f"Opslaan mislukt: {e}")

        with c2:
            del_ids = edited.loc[edited["Selecteer"] == True, "ID"].tolist()

            if st.button("🗑️ Verwijder geselecteerde producten", use_container_width=True):
                if not del_ids:
                    st.warning("Selecteer eerst één of meer producten.")
                else:
                    st.session_state.products = st.session_state.products[
                        ~st.session_state.products["id"].isin(del_ids)
                    ]

                    save_data()
                    st.success(f"Verwijderd: {del_ids}")
                    st.rerun()

    with st.expander("🛠️ Reparatie / import-check voor products.csv (GitHub)"):
        st.info("Hier kun je het productbestand controleren of repareren als import mislukt is.")
        st.markdown("*(Alleen zichtbaar op de pagina ‘Producten’)*")
